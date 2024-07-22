from DrissionPage import ChromiumPage
from lxml import etree
import requests
from openpyxl import load_workbook
import os
from tqdm import tqdm

keywords = []  # 定义一个列表存放关键词，我是需要艺术画加上不同关键字所以这样子设置
wb = load_workbook('素材库搜索模版.xlsx')  # 读取excel中的数据存入keywords
ws = wb['Sheet1']  # 操作工作表1
for i in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
    key = ws[f'{i}1'].value  # 获取这个excel中的第一
    if key is not None:
        # print(key)
        keywords.append(key)
    else:
        pass
# print(keywords)
# 指定要创建的文件夹路径
folder_path = r'C:\Users\Administrator\Desktop\Amazon_img'  # 自定义路径，按照keywords在这个路径下创建不同的文件夹存放图片

drive = ChromiumPage()

drive.get(r'https://www.amazon.com/')  # 跳转到亚马逊
drive.wait(2)
# keywords = ['logo', 'gold']  # 输入关键词，我是需要艺术画加上不同关键字所以这样子设置

for name in keywords:
    two_parts_name = None
    three_parts_name = None
    # 对关键词进行判断,如果带空格的关键词一个进行分割然后中间加上加号去赋值给需要监听的数据包
    if ' ' in name:
        parts = name.split()
        # 检查是否至少有两个部分
        if len(parts) == 2:
            # 将分割后的两个部分分别赋值给两个变量
            name1 = parts[0]
            name2 = parts[1]
            two_parts_name = name1 + '+' + name2
            print(two_parts_name)
        if len(parts) == 3:
            # 将分割后的两个部分分别赋值给两个变量
            name1 = parts[0]
            name2 = parts[1]
            name3 = parts[2]
            three_parts_name = name1 + '+' + name2 + '+' + name3
            print(three_parts_name)
    # 如果没有空格就清空原来的值以防后面判断出错误
    else:
        two_parts_name = None
        three_parts_name = None
    drive.ele('@id=twotabsearchtextbox').clear()  # 清除文本框内容
    drive.ele('@id=twotabsearchtextbox').input(f'wall art {name}\n')  # 输入内容到文本框中
    drive.ele('@id=a-autoid-0').click()  # 对排序进行切换,切换到销量最高
    drive.ele('@id=s-result-sort-select_5').click()
    x = 1  # 定义一个累加器对爬取的图片进行排序
    # 销量最高的前三页有价值所以这样子设置,如果需要更多页就增加一个循环,但是可能会因为这个关键词的商品页数小于你的循环次数而报错
    for i in range(1, 4):
        # s?k=后面修改成自己定义的关键词比较好抓到包,这里是我自己需要抓这个所以没修改
        # 对需要监听的数据包进行判断，如果是中间带空格的关键词会抓取不到需要进行处理
        if two_parts_name:
            drive.listen.start(f's?k=wall+art+{two_parts_name}&s=exact-aware-popularity-rank&')
        elif three_parts_name:
            drive.listen.start(f's?k=wall+art+{three_parts_name}&s=exact-aware-popularity-rank&')
        else:
            drive.listen.start(f's?k=wall+art+{name}&s=exact-aware-popularity-rank&')

        drive.refresh()  # 刷新网页,不然亚马逊不加载包出来
        resp = drive.listen.wait()  # 等待一个数据包的加载
        text = resp.response.body  # 获取数据包的响应对象
        # 将HTML内容转换为etree对象
        tree = etree.HTML(text)
        # 使用XPath查找class为s-image的img标签的src属性
        # 注意：XPath中class属性需要使用@class并且可能需要处理空格（如果有的话）
        # 这里我们假设class值中没有空格
        image_urls = tree.xpath('//img[@class="s-images"]/@src')
        # 进行图片保存,不同的关键词可以在桌面创建不一样的文件夹
        path = os.path.join(folder_path, name)  # 拼接保存路径
        # 对文件夹进行一个是否存在的判断
        if not os.path.exists(path):
            os.makedirs(path)  # 创建文件夹,路径在第十一行自定义

        for url in tqdm(image_urls, desc=f'正在爬取关键词{name}第{i}页'):
            if url == r'https://m.media-amazon.com/images/I/111mHoVK0kL._SS200_.png':
                # https://m.media-amazon.com/images/I/111mHoVK0kL._SS200_.png 意外获取到的图片链接
                pass
            else:
                with open(f'{path}/{name}{x}.jpg', 'wb') as f:
                    f.write(requests.get(url).content)  # 使用requests对获得的链接进行爬取并转码保存
                x += 1
        print(f'爬取完成关键词{name}第{i}页')  # 一个日志输出
        drive.ele('@text()=Next').click()  # 切换到下一页,就是切换下一页
        drive.wait(3)
        drive.refresh()  # 再次刷新以防下一次循环没有加载出数据包
